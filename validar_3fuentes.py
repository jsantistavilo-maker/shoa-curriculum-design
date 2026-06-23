"""Validación cruzada de horas entre 3 fuentes: PDF, Excel, Word."""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

BASE_DIR = Path(__file__).resolve().parent
PDF_PATH = BASE_DIR / "tabla cruzasa SHOA.pdf"
XLSX_WORK = BASE_DIR / "Tabla_resumen_work.xlsx"
PROGRAMAS_JSON = BASE_DIR / "programas_estructurados.json"
OUTPUT_XLSX = BASE_DIR / "validacion_3fuentes.xlsx"

CODIGOS_CANON = [
    "ABM101", "GDM301", "GP204", "HDP302", "HO202", "LR303",
    "LSM102", "MGG103", "OA203", "PH401", "PhyOce104", "RS201",
    "SG304", "WTC205",
]


def _nc(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())


CANON_MAP = {_nc(c): c for c in CODIGOS_CANON}


def canon(raw: str) -> str | None:
    key = _nc(raw)
    for norm, code in CANON_MAP.items():
        if norm in key or key in norm:
            return code
    return None


# ═══════════════════════════════════════════════════════════════════════════
# 1. PDF — tabla cruzada SHOA
# ═══════════════════════════════════════════════════════════════════════════
def extraer_pdf() -> dict[str, dict[str, float]]:
    pdf = pdfplumber.open(str(PDF_PATH))
    asig_re = re.compile(r"([A-Za-z]+)\s*(\d{3})")
    horas: dict[str, dict[str, float]] = defaultdict(lambda: {"T": 0.0, "P": 0.0, "SG": 0.0})

    for page in pdf.pages:
        tables = page.extract_tables()
        for table in tables:
            for row in table:
                if not row or len(row) < 7:
                    continue
                module_cell = row[3] or ""
                t_cell = row[4] or ""
                p_cell = row[5] or ""
                sg_cell = row[6] or ""

                if "TOTAL" in module_cell.upper():
                    continue

                modules = module_cell.split("\n")
                ts = t_cell.split("\n")
                ps = p_cell.split("\n")
                sgs = sg_cell.split("\n")

                for idx, mod in enumerate(modules):
                    mod = mod.strip()
                    if not mod:
                        continue
                    m = asig_re.search(mod)
                    if not m:
                        continue
                    raw_code = f"{m.group(1)} {m.group(2)}"
                    code = canon(raw_code)
                    if not code:
                        continue

                    def safe_num(lst, i):
                        if i >= len(lst):
                            return 0.0
                        v = lst[i].strip()
                        try:
                            return float(v)
                        except (ValueError, TypeError):
                            return 0.0

                    horas[code]["T"] += safe_num(ts, idx)
                    horas[code]["P"] += safe_num(ps, idx)
                    horas[code]["SG"] += safe_num(sgs, idx)

    pdf.close()
    return dict(horas)


# ═══════════════════════════════════════════════════════════════════════════
# 2. Excel — Horas_Asignaturas_SHOA
# ═══════════════════════════════════════════════════════════════════════════
def extraer_excel() -> dict[str, dict[str, float]]:
    from openpyxl import load_workbook

    wb = load_workbook(str(XLSX_WORK), data_only=True)
    ws = wb["Horas_Asignaturas_SHOA"]
    asig_re = re.compile(r"([A-Za-z]+)\s*(\d{3})")
    horas: dict[str, dict[str, float]] = defaultdict(lambda: {"T": 0.0, "P": 0.0, "SG": 0.0})

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        val_a = str(row[0].value or "").strip()
        if not val_a:
            continue
        m = asig_re.search(val_a)
        if not m:
            continue
        raw_code = f"{m.group(1)} {m.group(2)}"
        code = canon(raw_code)
        if not code:
            continue

        def safe(v):
            try:
                return float(v or 0)
            except (ValueError, TypeError):
                return 0.0

        horas[code]["T"] += safe(row[2].value)
        horas[code]["P"] += safe(row[3].value)
        horas[code]["SG"] += safe(row[4].value)

    return dict(horas)


# ═══════════════════════════════════════════════════════════════════════════
# 3. Word — programas_estructurados.json
# ═══════════════════════════════════════════════════════════════════════════
def extraer_word() -> dict[str, dict[str, float]]:
    data = json.loads(PROGRAMAS_JSON.read_text(encoding="utf-8"))
    horas: dict[str, dict[str, float]] = {}
    for p in data.get("programas", []):
        cod = p.get("codigo_asignatura")
        if not cod or p.get("error"):
            continue
        h = p.get("horas", {})
        horas[cod] = {
            "T": float(h.get("T", 0)),
            "P": float(h.get("P", 0)),
            "SG": float(h.get("SG", 0)),
        }
    return horas


# ═══════════════════════════════════════════════════════════════════════════
# 4. Comparación y clasificación
# ═══════════════════════════════════════════════════════════════════════════
def comparar(
    pdf_h: dict, excel_h: dict, word_h: dict
) -> list[dict]:
    resultados = []
    for cod in CODIGOS_CANON:
        ph = pdf_h.get(cod, {"T": 0, "P": 0, "SG": 0})
        eh = excel_h.get(cod, {"T": 0, "P": 0, "SG": 0})
        wh = word_h.get(cod, {"T": 0, "P": 0, "SG": 0})

        row = {"codigo": cod}
        max_diff = 0
        fuente_verdad: dict[str, float] = {}
        notas: list[str] = []

        for campo, label in [("T", "T"), ("P", "P"), ("SG", "AE")]:
            vp = int(ph.get(campo, 0))
            ve = int(eh.get(campo, 0))
            vw = int(wh.get(campo, 0))
            row[f"{label}_PDF"] = vp
            row[f"{label}_Excel"] = ve
            row[f"{label}_Word"] = vw

            vals = [vp, ve, vw]
            diff = max(vals) - min(vals)
            row[f"diff_{label}"] = diff
            max_diff = max(max_diff, diff)

            if vp == ve == vw:
                row[f"ok_{label}"] = True
                fuente_verdad[campo] = vp
            elif vp == ve and vw != vp:
                row[f"ok_{label}"] = False
                fuente_verdad[campo] = vp
                notas.append(f"{label}: PDF=Excel={vp}, Word={vw} -> usar PDF/Excel")
            elif vp == vw and ve != vp:
                row[f"ok_{label}"] = False
                fuente_verdad[campo] = vp
                notas.append(f"{label}: PDF=Word={vp}, Excel={ve} -> usar PDF/Word")
            elif ve == vw and vp != ve:
                row[f"ok_{label}"] = False
                fuente_verdad[campo] = ve
                notas.append(f"{label}: Excel=Word={ve}, PDF={vp} -> usar Excel/Word")
            else:
                row[f"ok_{label}"] = False
                fuente_verdad[campo] = ve
                notas.append(f"{label}: 3 difieren (PDF={vp}, Excel={ve}, Word={vw}) -> revision manual")

        tot_pdf = row["T_PDF"] + row["P_PDF"] + row["AE_PDF"]
        tot_excel = row["T_Excel"] + row["P_Excel"] + row["AE_Excel"]
        tot_word = row["T_Word"] + row["P_Word"] + row["AE_Word"]
        row["Total_PDF"] = tot_pdf
        row["Total_Excel"] = tot_excel
        row["Total_Word"] = tot_word
        row["ok_Total"] = (tot_pdf == tot_excel == tot_word)

        if max_diff == 0:
            row["estado"] = "VALIDADA"
            row["emoji"] = "✅"
        elif max_diff <= 3:
            row["estado"] = "DISCREPANCIA LEVE"
            row["emoji"] = "⚠️"
        else:
            row["estado"] = "DISCREPANCIA GRAVE"
            row["emoji"] = "❌"

        row["fuente_verdad"] = fuente_verdad
        row["notas"] = "; ".join(notas) if notas else "Las 3 fuentes coinciden"
        resultados.append(row)

    return resultados


# ═══════════════════════════════════════════════════════════════════════════
# 5. Generar Excel de validación
# ═══════════════════════════════════════════════════════════════════════════
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD = Font(bold=True)


def generar_xlsx(resultados: list[dict]) -> None:
    wb = Workbook()

    # ── Hoja 1: Resumen ejecutivo ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    validadas = sum(1 for r in resultados if r["estado"] == "VALIDADA")
    leves = sum(1 for r in resultados if r["estado"] == "DISCREPANCIA LEVE")
    graves = sum(1 for r in resultados if r["estado"] == "DISCREPANCIA GRAVE")

    ws1.append(["Validación de Horas — 3 Fuentes (PDF, Excel, Word)"])
    ws1.merge_cells("A1:D1")
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.append([])
    ws1.append(["Indicador", "Cantidad", "Porcentaje"])
    ws1.append(["✅ Asignaturas validadas", validadas, f"{validadas/14*100:.0f}%"])
    ws1.append(["⚠️ Discrepancia leve (1-3h)", leves, f"{leves/14*100:.0f}%"])
    ws1.append(["❌ Discrepancia grave (>3h)", graves, f"{graves/14*100:.0f}%"])
    ws1.append(["Total asignaturas", 14, "100%"])
    ws1.append([])
    ws1.append(["Fuentes comparadas:"])
    ws1.append(["  1. PDF: tabla cruzasa SHOA.pdf"])
    ws1.append(["  2. Excel: Tabla resumen.xlsx -> Horas_Asignaturas_SHOA"])
    ws1.append(["  3. Word: programas_estructurados.json (fila TOTAL)"])
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 12

    for row in ws1.iter_rows(min_row=4, max_row=7, min_col=1, max_col=1):
        for cell in row:
            if "validadas" in str(cell.value or ""):
                cell.fill = GREEN_FILL
            elif "leve" in str(cell.value or ""):
                cell.fill = YELLOW_FILL
            elif "grave" in str(cell.value or ""):
                cell.fill = RED_FILL

    # ── Hoja 2: Tabla comparativa completa ───────────────────────────────
    ws2 = wb.create_sheet("Comparativa Completa")
    headers = [
        "Asignatura", "Estado",
        "T PDF", "T Excel", "T Word", "Diff T",
        "P PDF", "P Excel", "P Word", "Diff P",
        "AE PDF", "AE Excel", "AE Word", "Diff AE",
        "Total PDF", "Total Excel", "Total Word",
        "Notas",
    ]
    ws2.append(headers)
    for ci in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=ci)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for r in resultados:
        row_data = [
            r["codigo"], f'{r["emoji"]} {r["estado"]}',
            r["T_PDF"], r["T_Excel"], r["T_Word"], r["diff_T"],
            r["P_PDF"], r["P_Excel"], r["P_Word"], r["diff_P"],
            r["AE_PDF"], r["AE_Excel"], r["AE_Word"], r["diff_AE"],
            r["Total_PDF"], r["Total_Excel"], r["Total_Word"],
            r["notas"],
        ]
        ws2.append(row_data)
        row_num = ws2.max_row
        fill = GREEN_FILL if r["estado"] == "VALIDADA" else (YELLOW_FILL if "LEVE" in r["estado"] else RED_FILL)
        ws2.cell(row=row_num, column=2).fill = fill
        for ci in [6, 10, 14]:
            c = ws2.cell(row=row_num, column=ci)
            if c.value and c.value > 0:
                c.fill = RED_FILL if c.value > 3 else YELLOW_FILL

    for ci in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 12
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["R"].width = 55

    # ── Hoja 3: Solo discrepancias ───────────────────────────────────────
    ws3 = wb.create_sheet("Discrepancias")
    disc_headers = [
        "Asignatura", "Estado",
        "T PDF", "T Excel", "T Word",
        "P PDF", "P Excel", "P Word",
        "AE PDF", "AE Excel", "AE Word",
        "Total PDF", "Total Excel", "Total Word",
        "Fuente de verdad T", "Fuente de verdad P", "Fuente de verdad AE",
        "Acción recomendada",
    ]
    ws3.append(disc_headers)
    for ci in range(1, len(disc_headers) + 1):
        cell = ws3.cell(row=1, column=ci)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    discs = [r for r in resultados if r["estado"] != "VALIDADA"]
    for r in discs:
        fv = r["fuente_verdad"]
        accion = "Corregir automáticamente" if "LEVE" in r["estado"] else "Revisión manual requerida"
        row_data = [
            r["codigo"], f'{r["emoji"]} {r["estado"]}',
            r["T_PDF"], r["T_Excel"], r["T_Word"],
            r["P_PDF"], r["P_Excel"], r["P_Word"],
            r["AE_PDF"], r["AE_Excel"], r["AE_Word"],
            r["Total_PDF"], r["Total_Excel"], r["Total_Word"],
            int(fv.get("T", 0)), int(fv.get("P", 0)), int(fv.get("SG", 0)),
            accion,
        ]
        ws3.append(row_data)
        row_num = ws3.max_row
        fill = YELLOW_FILL if "LEVE" in r["estado"] else RED_FILL
        ws3.cell(row=row_num, column=2).fill = fill

    for ci in range(1, len(disc_headers) + 1):
        ws3.column_dimensions[get_column_letter(ci)].width = 14
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["R"].width = 30

    wb.save(str(OUTPUT_XLSX))


# ═══════════════════════════════════════════════════════════════════════════
# 6. Main
# ═══════════════════════════════════════════════════════════════════════════
def main() -> None:
    print("=" * 80)
    print("  VALIDACIÓN CRUZADA DE HORAS — 3 FUENTES")
    print("=" * 80)

    print("\n📄 Extrayendo horas del PDF...")
    pdf_h = extraer_pdf()
    print(f"   {len(pdf_h)} asignaturas encontradas en PDF")

    print("📊 Extrayendo horas del Excel...")
    excel_h = extraer_excel()
    print(f"   {len(excel_h)} asignaturas encontradas en Excel")

    print("📝 Extrayendo horas del Word...")
    word_h = extraer_word()
    print(f"   {len(word_h)} asignaturas encontradas en Word")

    print("\n🔍 Comparando las 3 fuentes...\n")
    resultados = comparar(pdf_h, excel_h, word_h)

    # ── Tabla consola ────────────────────────────────────────────────────
    hdr = (
        f"{'Asig':12s} | {'T PDF':>5s} {'T Xls':>5s} {'T Wrd':>5s} |"
        f" {'P PDF':>5s} {'P Xls':>5s} {'P Wrd':>5s} |"
        f" {'AE PDF':>6s} {'AE Xls':>6s} {'AE Wrd':>6s} |"
        f" {'TotP':>4s} {'TotX':>4s} {'TotW':>4s} | Estado"
    )
    sep = "-" * len(hdr)
    print(hdr)
    print(sep)

    for r in resultados:
        print(
            f"{r['codigo']:12s} |"
            f" {r['T_PDF']:>5d} {r['T_Excel']:>5d} {r['T_Word']:>5d} |"
            f" {r['P_PDF']:>5d} {r['P_Excel']:>5d} {r['P_Word']:>5d} |"
            f" {r['AE_PDF']:>6d} {r['AE_Excel']:>6d} {r['AE_Word']:>6d} |"
            f" {r['Total_PDF']:>4d} {r['Total_Excel']:>4d} {r['Total_Word']:>4d} |"
            f" {r['emoji']} {r['estado']}"
        )

    print(sep)

    validadas = sum(1 for r in resultados if r["estado"] == "VALIDADA")
    leves = sum(1 for r in resultados if r["estado"] == "DISCREPANCIA LEVE")
    graves = sum(1 for r in resultados if r["estado"] == "DISCREPANCIA GRAVE")

    # ── Generar Excel ────────────────────────────────────────────────────
    print(f"\n📊 Generando {OUTPUT_XLSX.name}...")
    generar_xlsx(resultados)

    # ── Actualizar programas_estructurados.json (solo leves) ─────────────
    actualizados = 0
    if leves > 0:
        prog_data = json.loads(PROGRAMAS_JSON.read_text(encoding="utf-8"))
        for r in resultados:
            if r["estado"] != "DISCREPANCIA LEVE":
                continue
            fv = r["fuente_verdad"]
            for p in prog_data.get("programas", []):
                if p.get("codigo_asignatura") == r["codigo"]:
                    p["horas"]["T"] = int(fv["T"])
                    p["horas"]["P"] = int(fv["P"])
                    p["horas"]["SG"] = int(fv["SG"])
                    p["horas"]["total"] = int(fv["T"] + fv["P"] + fv["SG"])
                    actualizados += 1
                    break
        PROGRAMAS_JSON.write_text(
            json.dumps(prog_data, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    # ── Resumen final ────────────────────────────────────────────────────
    print("\n" + "=" * 80)
    print(f"  ✅ {validadas}/14 asignaturas validadas correctamente")
    if leves:
        print(f"  ⚠️  {leves} con discrepancia leve → corregidas automáticamente")
    if graves:
        print(f"  ❌ {graves} con discrepancia grave → requieren revisión manual")
    if actualizados:
        print(f"  📝 {actualizados} asignaturas actualizadas en programas_estructurados.json")
    print(f"  📋 Ver {OUTPUT_XLSX.name} para detalle completo")
    print("=" * 80)

    # Detalle de discrepancias
    for r in resultados:
        if r["estado"] != "VALIDADA":
            print(f"\n  {r['emoji']} {r['codigo']}:")
            print(f"     {r['notas']}")


if __name__ == "__main__":
    main()
